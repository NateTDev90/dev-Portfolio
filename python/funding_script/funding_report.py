import pandas as pd
import glob
import os
from datetime import datetime
import csv
from openpyxl import load_workbook

# Configuration Section
WORKING_DIRECTORY = r"S:\IT\Operations\Online Account Opening\FundingReports\Python Funding Script"  # Set your working directory here
REPORT_OUTPUT_DIRECTORY = r"S:\IT\Operations\Online Account Opening\FundingReports\FundingReports"  # Set your report output directory here
CSV_FILE_PATTERN = "*.csv"  # Pattern to find CSV files
EXCEL_FILE = "Transaction_Listing.xlsx"  # Name of the Excel file
SHEET_NAME = "Report (Page 1)"  # Name of the sheet in the Excel file

# Ensure the report output directory exists
os.makedirs(REPORT_OUTPUT_DIRECTORY, exist_ok=True)

# Set the working directory
os.chdir(WORKING_DIRECTORY)

# Find the CSV file
csv_files = glob.glob(CSV_FILE_PATTERN)
if not csv_files:
    raise FileNotFoundError("No CSV file found in the directory.")
csv_file = csv_files[0]
print(f"Processing CSV file: {csv_file}")

# Read the Excel file
try:
    df_xlsx = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
except FileNotFoundError:
    raise FileNotFoundError(f"{EXCEL_FILE} not found.")
except ValueError:
    raise ValueError(f"Sheet '{SHEET_NAME}' not found.")

# Clean Excel data
df_xlsx.columns = [col.replace('.', '') for col in df_xlsx.columns]
required_xlsx_cols = ['Invoice Number']
for col in required_xlsx_cols:
    if col not in df_xlsx.columns:
        raise KeyError(f"'{col}' not found in Excel. Available: {list(df_xlsx.columns)}")
df_xlsx['Invoice Number'] = pd.to_numeric(df_xlsx['Invoice Number'], errors='coerce').astype(str)

# Parse the CSV file
csv_data = []
with open(csv_file, newline='', encoding='utf-8') as f:
    reader = csv.reader(f, quotechar='"', quoting=csv.QUOTE_MINIMAL)
    csv_headers = next(reader)
    csv_headers = [header for header in csv_headers if header.strip()]
    expected_cols = len(csv_headers)
    for row in reader:
        if len(row) > expected_cols:
            row = row[:expected_cols]
        elif len(row) < expected_cols:
            row.extend([''] * (expected_cols - len(row)))
        csv_data.append(row)

# Create CSV DataFrame
df_csv = pd.DataFrame(csv_data, columns=csv_headers)

# Verify CSV columns
required_csv_cols = ['APPLICATION NUMBER', 'FIRSTNAME', 'LASTNAME', 'FUNDINGAMOUNT', 'ACCTNUMBER', 'PRODNAME', 'RELATIONSHIP']
for col in required_csv_cols:
    if col not in df_csv.columns:
        if col == 'RELATIONSHIP':
            print(f"Warning: '{col}' column not found in CSV. All records will be included.")
            df_csv['RELATIONSHIP'] = ''  # Add empty RELATIONSHIP column if it doesn't exist
        else:
            raise KeyError(f"'{col}' not found in CSV. Available: {list(df_csv.columns)}")

# Convert data types
df_csv['APPLICATION NUMBER'] = df_csv['APPLICATION NUMBER'].astype(str)
df_csv['FUNDINGAMOUNT'] = pd.to_numeric(df_csv['FUNDINGAMOUNT'], errors='coerce').fillna(0).astype(float)
df_csv['ACCTNUMBER'] = df_csv['ACCTNUMBER'].astype(str)

# Prepare the final report
final_columns = ['FIRSTNAME', 'LASTNAME', 'FUNDINGAMOUNT', 'ACCTNUMBER', 'PRODNAME', 'APPLICATION NUMBER']
final_report = []

# Match records between Excel and CSV, excluding joint owners
for idx, xlsx_row in df_xlsx.iterrows():
    invoice_num = xlsx_row['Invoice Number']
    matching_rows = df_csv[df_csv['APPLICATION NUMBER'] == invoice_num]
    
    for _, csv_row in matching_rows.iterrows():
        # Skip joint owners (check if RELATIONSHIP column indicates a joint owner)
        relationship = str(csv_row.get('RELATIONSHIP', '')).strip().upper()
        if 'JOINT' in relationship:
            continue  # Skip this record as it's a joint owner
            
        report_row = {
            'FIRSTNAME': str(csv_row['FIRSTNAME']),
            'LASTNAME': str(csv_row['LASTNAME']),
            'FUNDINGAMOUNT': csv_row['FUNDINGAMOUNT'],
            'ACCTNUMBER': str(csv_row['ACCTNUMBER']),
            'PRODNAME': str(csv_row['PRODNAME']),
            'APPLICATION NUMBER': str(csv_row['APPLICATION NUMBER'])
        }
        final_report.append(report_row)

# Create final DataFrame
df_final = pd.DataFrame(final_report, columns=final_columns)

# Generate output filename with current date
current_date = datetime.now().strftime("%m%d%y")
output_file = os.path.join(REPORT_OUTPUT_DIRECTORY, f"MbrFundingReport {current_date}.xlsx")

# Save the report to Excel
df_final.to_excel(output_file, index=False)

# Load the workbook to apply formatting
wb = load_workbook(output_file)
ws = wb.active

# Auto-fit column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Apply currency formatting to 'FUNDINGAMOUNT' column
funding_col_idx = final_columns.index('FUNDINGAMOUNT') + 1  # Excel columns start at 1
for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
    cell = ws.cell(row=row, column=funding_col_idx)
    cell.number_format = '$#,##0.00'

# Save the formatted workbook
wb.save(output_file)

print(f"Final report generated: {output_file}")
print(f"Number of records in final report: {len(df_final)}")
print("Note: Joint owners have been excluded from the report.")
